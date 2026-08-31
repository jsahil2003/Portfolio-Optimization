"""
Builds the submission workbook required by the challenge:
composition & weights, returns, MDD, benchmark comparison, and model
logic/assumptions - all in one .xlsx, per the "Submission Requirements".

A formatted "Summary" cover sheet (composition, returns, MDD, benchmark
comparison, assumptions, NAV chart) leads the workbook; the remaining
sheets hold the full supporting detail (every rebalance's weights, daily
NAV/returns, full drawdown series, complete trade log).
"""
import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F3864")
SECTION_FONT = Font(size=12, bold=True, color="1F3864")


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autofit(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def build_workbook(path: str, weights_by_date: dict, nav: pd.Series,
                    benchmark_nav: pd.Series, report: dict, trades: pd.DataFrame,
                    benchmark_navs: dict = None, starting_capital: float = 1_00_00_000):
    """benchmark_nav: single legacy benchmark (kept for backward compatibility).
    benchmark_navs: {name: nav_series} for one or more named benchmarks,
    each added as its own column in the Returns sheet."""
    all_benchmarks = dict(benchmark_navs or {})
    if benchmark_nav is not None:
        all_benchmarks.setdefault("benchmark", benchmark_nav)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        comp_rows = []
        for date, weights in weights_by_date.items():
            for ticker, w in weights.items():
                comp_rows.append({"rebalance_date": date, "ticker": ticker, "weight": w})
        comp_df = pd.DataFrame(comp_rows)
        comp_df.to_excel(writer, sheet_name="Composition_Weights", index=False)

        nav_df = pd.DataFrame({"date": nav.index, "portfolio_nav": nav.values})
        nav_df["daily_return"] = nav_df["portfolio_nav"].pct_change()
        for name, bench_nav in all_benchmarks.items():
            nav_df = nav_df.merge(
                pd.DataFrame({"date": bench_nav.index, f"{name}_nav": bench_nav.values}),
                on="date", how="left",
            )
        nav_df.to_excel(writer, sheet_name="Returns", index=False)

        running_max = nav.cummax()
        dd = (nav / running_max - 1.0)
        dd_df = pd.DataFrame({"date": nav.index, "nav": nav.values,
                               "running_max": running_max.values, "drawdown": dd.values})
        dd_df.to_excel(writer, sheet_name="Drawdown", index=False)

        summary_rows = [{"metric": k, "value": v} for k, v in report.items()
                         if not isinstance(v, pd.Series)]
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary_Metrics", index=False)

        if not trades.empty:
            trades.to_excel(writer, sheet_name="Trade_Log", index=False)

        assumption_labels = [
            "Universe", "Max stocks", "Backtest period", "Transaction cost",
            "Starting capital", "Rebalance frequency", "Weighting scheme",
            "Selection method", "Composite factor weights", "Turnover buffer",
            "Sector cap", "Share trading", "Lookahead bias", "Risk-free rate", "Benchmarks",
        ]
        assumption_values = [
            "Nifty 100 + Nifty Midcap 100 + Nifty Smallcap 100 (live pull, niftyindices.com)",
            10, "2021-01-01 to 2025-12-31", "0.1% per transaction",
            "INR 1,00,00,000", "Monthly",
            "Ledoit-Wolf shrinkage minimum-variance, capped at 20% per name",
            "Top 10 by composite of momentum, low-vol, Parkinson intraday vol, 52-week-high, "
            "liquidity z-scores",
            "momentum 0.40 / low_vol 0.25 / parkinson_vol 0.10 / high52 0.15 / liquidity 0.10 "
            "(walk-forward validated: tuned on 2021-2023, confirmed on unseen 2024-2025, "
            "re-validated with parkinson_vol added and stress-tested)",
            "A held stock stays as long as its rank is within top 13 (10+3), only new entrants "
            "need top-10; buffer=3 chosen via walk-forward grid {0,2,3,5,8}, improving Sharpe/"
            "return/drawdown/turnover on train and test independently",
            "Max 35% of portfolio weight per NSE industry, tested via backtest: removed every "
            ">40%-in-one-sector rebalance (22/60 -> 0/60) while improving return/drawdown/Sharpe",
            "Whole shares only - fractional trades are not permitted; rounding shortfall held as cash",
            "None - all five factors built from historical price/volume data only, "
            "genuinely point-in-time safe (no fundamentals snapshot used)",
            "0%",
            ", ".join(all_benchmarks.keys()) if all_benchmarks else "none",
        ]
        assumptions = pd.DataFrame({"assumption": assumption_labels, "value": assumption_values})
        assumptions.to_excel(writer, sheet_name="Model_Logic_Assumptions", index=False)

        # ---- Summary cover sheet ----
        wb = writer.book
        summary_ws = wb.create_sheet("Summary", 0)

        latest_date = max(weights_by_date.keys())
        latest_weights = weights_by_date[latest_date].sort_values(ascending=False)
        final_value = nav.iloc[-1]
        total_return_pct = final_value / starting_capital - 1.0

        row = 1
        summary_ws.cell(row=row, column=1, value="Finesse x Citadel Portfolio Challenge — Submission Summary")
        summary_ws.cell(row=row, column=1).font = TITLE_FONT
        row += 2

        summary_ws.cell(row=row, column=1, value="1. Portfolio Composition & Weights")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        summary_ws.cell(row=row, column=1, value=f"(latest rebalance: {latest_date.date()})")
        row += 1
        header_row = row
        summary_ws.cell(row=row, column=1, value="Ticker")
        summary_ws.cell(row=row, column=2, value="Weight")
        row += 1
        for ticker, w in latest_weights.items():
            summary_ws.cell(row=row, column=1, value=ticker)
            c = summary_ws.cell(row=row, column=2, value=w)
            c.number_format = "0.00%"
            row += 1
        for cell in summary_ws[header_row]:
            if cell.value is not None:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        row += 1

        summary_ws.cell(row=row, column=1, value="2. Returns")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        return_rows = [
            ("Starting capital (INR)", starting_capital, "#,##0"),
            ("Final portfolio value (INR)", final_value, "#,##0"),
            ("Total net PNL (INR)", report["total_net_pnl"], "#,##0"),
            ("Total return (absolute, %)", total_return_pct, "0.00%"),
            ("Annualized return (%)", report["annualized_return"], "0.00%"),
        ]
        for label, value, fmt in return_rows:
            summary_ws.cell(row=row, column=1, value=label)
            c = summary_ws.cell(row=row, column=2, value=value)
            c.number_format = fmt
            row += 1
        row += 1

        summary_ws.cell(row=row, column=1, value="3. Maximum Drawdown")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        summary_ws.cell(row=row, column=1, value="Max drawdown (%)")
        c = summary_ws.cell(row=row, column=2, value=report["max_drawdown"])
        c.number_format = "0.00%"
        row += 2

        summary_ws.cell(row=row, column=1, value="4. Benchmark Comparison")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        header_row = row
        summary_ws.cell(row=row, column=1, value="Metric")
        summary_ws.cell(row=row, column=2, value="This Strategy")
        col = 3
        for name in all_benchmarks:
            summary_ws.cell(row=row, column=col, value=name)
            col += 1
        row += 1
        bench_metric_rows = [
            ("Annualized return", "annualized_return", "0.00%"),
            ("Max drawdown", "max_drawdown", "0.00%"),
            ("Sharpe ratio", "sharpe_ratio", "0.00"),
        ]
        for label, key, fmt in bench_metric_rows:
            summary_ws.cell(row=row, column=1, value=label)
            c = summary_ws.cell(row=row, column=2, value=report[key])
            c.number_format = fmt
            col = 3
            for name in all_benchmarks:
                bench_key = f"{name}_{key}"
                c = summary_ws.cell(row=row, column=col, value=report.get(bench_key))
                c.number_format = fmt
                col += 1
            row += 1
        summary_ws.cell(row=row, column=1, value="Excess annualized return vs.")
        for i, name in enumerate(all_benchmarks):
            c = summary_ws.cell(row=row, column=3 + i, value=report.get(f"{name}_excess_annualized_return"))
            c.number_format = "0.00%"
        row += 1
        for cell in summary_ws[header_row]:
            if cell.value is not None:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        row += 1

        summary_ws.cell(row=row, column=1, value="5. Secondary Metrics")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        secondary_rows = [
            ("Gain-to-loss ratio", report.get("gain_to_loss_ratio"), "0.00"),
            ("Accuracy (win rate)", report.get("accuracy"), "0.00%"),
            ("Total trades", report.get("total_trades"), "#,##0"),
            ("Avg turnover per rebalance", report.get("avg_turnover_per_rebalance"), "0.00%"),
            ("Total transaction cost (INR)", report.get("total_transaction_cost"), "#,##0"),
        ]
        for label, value, fmt in secondary_rows:
            summary_ws.cell(row=row, column=1, value=label)
            c = summary_ws.cell(row=row, column=2, value=value)
            c.number_format = fmt
            row += 1
        row += 1

        summary_ws.cell(row=row, column=1, value="6. Model Logic & Assumptions")
        summary_ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        for label, value in zip(assumption_labels, assumption_values):
            summary_ws.cell(row=row, column=1, value=label)
            summary_ws.cell(row=row, column=1).font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=str(value))
            summary_ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        summary_ws.column_dimensions["A"].width = 30
        summary_ws.column_dimensions["B"].width = 60
        for col_letter in ["C", "D", "E"]:
            summary_ws.column_dimensions[col_letter].width = 18

        # ---- NAV vs. benchmark chart, on the Summary sheet, sourced from Returns ----
        returns_ws = wb["Returns"]
        n_rows = returns_ws.max_row
        chart = LineChart()
        chart.title = "Portfolio NAV vs. Benchmarks"
        chart.y_axis.title = "Value (INR)"
        chart.x_axis.title = "Date"
        chart.height, chart.width = 10, 22

        nav_col = 2  # "date" is col A, "portfolio_nav" is col B
        chart.add_data(Reference(returns_ws, min_col=nav_col, min_row=1, max_row=n_rows), titles_from_data=True)
        col = 4  # column C is "daily_return", benchmark nav columns start at D
        for _ in all_benchmarks:
            chart.add_data(Reference(returns_ws, min_col=col, min_row=1, max_row=n_rows), titles_from_data=True)
            col += 1
        chart.set_categories(Reference(returns_ws, min_col=1, min_row=2, max_row=n_rows))
        summary_ws.add_chart(chart, f"D{header_row + 20}")

        # ---- Formatting pass on the detail sheets ----
        for sheet_name in ["Composition_Weights", "Returns", "Drawdown", "Summary_Metrics",
                            "Trade_Log", "Model_Logic_Assumptions"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            _style_header_row(ws)
            _autofit(ws)

        pct_cols = {
            "Composition_Weights": ["weight"],
            "Returns": ["daily_return"],
            "Drawdown": ["drawdown"],
        }
        for sheet_name, cols in pct_cols.items():
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]
            for col_name in cols:
                if col_name not in headers:
                    continue
                col_idx = headers.index(col_name) + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "0.00%"
